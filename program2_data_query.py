#!/usr/bin/env python3
"""
程式二：資料查詢與整理
目的：根據使用者指定的參數，從本地資料庫中提取數據，並疊加K線圖
"""

import os
import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Rectangle
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image
import logging
import argparse
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class StockDataQuery:
    """股權分佈資料查詢與整理系統"""
    
    def __init__(self, data_dir: str = "stock_data"):
        """
        初始化查詢系統
        
        Args:
            data_dir: 資料儲存目錄
        """
        self.data_dir = Path(data_dir)
        self.wearn_url = "https://stock.wearn.com/cdata.asp"
        
    def find_closest_date(self, stock_code: str, target_date: str, 
                         direction: str = "after") -> Optional[Tuple[str, bool]]:
        """
        尋找最接近的可用日期
        
        Args:
            stock_code: 股票代碼
            target_date: 目標日期 (YYYY-MM-DD)
            direction: 搜尋方向 ("after" 或 "before")
            
        Returns:
            (最接近的日期, 是否有警告)
        """
        stock_dir = self.data_dir / stock_code
        if not stock_dir.exists():
            logger.error(f"股票 {stock_code} 資料夾不存在")
            return None
            
        # 獲取所有可用日期
        available_dates = []
        for file_path in stock_dir.glob("*.json"):
            date_str = file_path.stem  # YYYY-MM-DD
            available_dates.append(date_str)
            
        if not available_dates:
            logger.error(f"股票 {stock_code} 無可用數據")
            return None
            
        available_dates.sort()
        target = datetime.strptime(target_date, "%Y-%m-%d")
        
        # 尋找大於等於目標日期的最近日期
        if direction == "after":
            for date_str in available_dates:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                if date_obj >= target:
                    return date_str, False
                    
            # 如果找不到，回退到小於目標日期的最近日期
            logger.warning(f"找不到 {target_date} 之後的數據，使用之前最近的日期")
            for date_str in reversed(available_dates):
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                if date_obj < target:
                    return date_str, True
                    
        # 尋找小於等於目標日期的最近日期
        else:
            for date_str in reversed(available_dates):
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                if date_obj <= target:
                    return date_str, False
                    
        return available_dates[0], True
        
    def load_stock_data(self, stock_code: str, start_date: str, end_date: str) -> pd.DataFrame:
        """
        載入指定期間的股權分佈數據
        
        Args:
            stock_code: 股票代碼
            start_date: 起始日期 (YYYY-MM-DD)
            end_date: 結束日期 (YYYY-MM-DD)
            
        Returns:
            股權分佈數據DataFrame
        """
        stock_dir = self.data_dir / stock_code
        if not stock_dir.exists():
            logger.error(f"股票 {stock_code} 資料夾不存在")
            return pd.DataFrame()
            
        # 尋找最接近的起始和結束日期
        actual_start, start_warning = self.find_closest_date(stock_code, start_date, "after") or (None, False)
        actual_end, end_warning = self.find_closest_date(stock_code, end_date, "before") or (None, False)
        
        if not actual_start or not actual_end:
            logger.error("無法找到有效的日期範圍")
            return pd.DataFrame()
            
        logger.info(f"實際查詢期間: {actual_start} 到 {actual_end}")
        
        if start_warning:
            logger.warning(f"起始日期 {start_date} 無可用數據，使用 {actual_start}")
        if end_warning:
            logger.warning(f"結束日期 {end_date} 無可用數據，使用 {actual_end}")
            
        # 載入期間內的所有數據
        data_list = []
        for file_path in sorted(stock_dir.glob("*.json")):
            date_str = file_path.stem
            if actual_start <= date_str <= actual_end:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    data['date_str'] = date_str
                    data_list.append(data)
                    
        return self.process_distribution_data(data_list)
        
    def process_distribution_data(self, data_list: List[Dict]) -> pd.DataFrame:
        """
        處理股權分佈數據
        
        Args:
            data_list: 原始數據列表
            
        Returns:
            處理後的DataFrame
        """
        if not data_list:
            return pd.DataFrame()
            
        processed_data = []
        for data in data_list:
            date_str = data['date_str']
            for item in data['distribution']:
                processed_data.append({
                    'date': date_str,
                    'level': item['level'],
                    'holders': int(item['holders']) if item['holders'] else 0,
                    'shares': int(item['shares']) if item['shares'] else 0,
                    'percentage': float(item['percentage'].replace('%', '')) if item['percentage'] else 0
                })
                
        df = pd.DataFrame(processed_data)
        df['date'] = pd.to_datetime(df['date'])
        return df
        
    def fetch_kline_data(self, stock_code: str, start_date: str, end_date: str) -> pd.DataFrame:
        """
        從Wearn.com獲取K線數據
        
        Args:
            stock_code: 股票代碼
            start_date: 起始日期 (YYYY-MM-DD)
            end_date: 結束日期 (YYYY-MM-DD)
            
        Returns:
            K線數據DataFrame
        """
        try:
            # 轉換日期格式
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            
            # 計算需要查詢的年月
            months_to_query = []
            current = start
            while current <= end:
                year = current.year - 1911  # 轉換為民國年
                month = current.month
                months_to_query.append((year, month))
                # 移動到下個月
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
                    
            kline_data = []
            for year, month in months_to_query:
                params = {
                    'Year': year,
                    'month': f"{month:02d}",
                    'kind': stock_code
                }
                
                response = requests.get(self.wearn_url, params=params, timeout=30)
                response.encoding = 'big5'
                
                # 解析HTML獲取K線數據
                soup = BeautifulSoup(response.text, 'html.parser')
                tables = soup.find_all('table')
                
                for table in tables:
                    rows = table.find_all('tr')
                    for row in rows[1:]:  # 跳過標題行
                        cols = row.find_all('td')
                        if len(cols) >= 6:
                            try:
                                date_text = cols[0].text.strip()
                                # 轉換民國年為西元年
                                if '/' in date_text:
                                    parts = date_text.split('/')
                                    if len(parts) == 3:
                                        west_year = int(parts[0]) + 1911
                                        date_str = f"{west_year}/{parts[1]}/{parts[2]}"
                                        
                                        kline_data.append({
                                            'date': pd.to_datetime(date_str),
                                            'open': float(cols[1].text.strip().replace(',', '')),
                                            'high': float(cols[2].text.strip().replace(',', '')),
                                            'low': float(cols[3].text.strip().replace(',', '')),
                                            'close': float(cols[4].text.strip().replace(',', '')),
                                            'volume': int(cols[5].text.strip().replace(',', ''))
                                        })
                            except (ValueError, IndexError) as e:
                                continue
                                
            if kline_data:
                df = pd.DataFrame(kline_data)
                df = df.sort_values('date')
                # 過濾日期範圍
                df = df[(df['date'] >= start_date) & (df['date'] <= end_date)]
                return df
            else:
                logger.warning("無法獲取K線數據")
                return pd.DataFrame()
                
        except Exception as e:
            logger.error(f"獲取K線數據失敗: {e}")
            return pd.DataFrame()
            
    def create_tables(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        建立三個分析表格
        
        Args:
            df: 股權分佈數據DataFrame
            
        Returns:
            包含三個表格的字典
        """
        tables = {}
        
        # 表格1: 人數
        pivot_holders = df.pivot_table(
            index='date',
            columns='level',
            values='holders',
            aggfunc='sum'
        )
        tables['holders'] = pivot_holders
        
        # 表格2: 股數/單位數
        pivot_shares = df.pivot_table(
            index='date',
            columns='level',
            values='shares',
            aggfunc='sum'
        )
        tables['shares'] = pivot_shares
        
        # 表格3: 占集保庫存數比例
        pivot_percentage = df.pivot_table(
            index='date',
            columns='level',
            values='percentage',
            aggfunc='sum'
        )
        tables['percentage'] = pivot_percentage
        
        return tables
        
    def create_chart_with_kline(self, table_data: pd.DataFrame, kline_data: pd.DataFrame,
                               title: str, ylabel: str) -> BytesIO:
        """
        建立疊加K線的圖表
        
        Args:
            table_data: 表格數據
            kline_data: K線數據
            title: 圖表標題
            ylabel: Y軸標籤
            
        Returns:
            圖表的BytesIO對象
        """
        fig, ax1 = plt.subplots(figsize=(15, 8))
        
        # 設定中文字體
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 繪製股權分佈數據
        colors = plt.cm.tab20(np.linspace(0, 1, len(table_data.columns)))
        for idx, col in enumerate(table_data.columns):
            ax1.plot(table_data.index, table_data[col], 
                    label=col, color=colors[idx], linewidth=2)
            
        ax1.set_xlabel('Date')
        ax1.set_ylabel(ylabel)
        ax1.set_title(title)
        ax1.legend(loc='upper left', bbox_to_anchor=(1.15, 1))
        ax1.grid(True, alpha=0.3)
        
        # 如果有K線數據，繪製在第二個Y軸
        if not kline_data.empty:
            ax2 = ax1.twinx()
            
            # 繪製K線圖
            for idx, row in kline_data.iterrows():
                date = mdates.date2num(row['date'])
                open_price = row['open']
                close_price = row['close']
                high_price = row['high']
                low_price = row['low']
                
                # 決定顏色
                color = 'red' if close_price >= open_price else 'green'
                
                # 繪製K線
                ax2.plot([date, date], [low_price, high_price], 
                        color=color, linewidth=0.5, alpha=0.6)
                
                # 繪製實體
                height = abs(close_price - open_price)
                bottom = min(close_price, open_price)
                rect = Rectangle((date - 0.3, bottom), 0.6, height,
                               facecolor=color, alpha=0.6)
                ax2.add_patch(rect)
                
            ax2.set_ylabel('Stock Price')
            ax2.grid(False)
            
            # 添加成交量
            ax3 = ax1.twinx()
            ax3.spines['right'].set_position(('outward', 60))
            ax3.bar(kline_data['date'], kline_data['volume'], 
                   alpha=0.3, color='gray', width=0.8)
            ax3.set_ylabel('Volume')
            ax3.grid(False)
            
        # 格式化X軸日期
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        ax1.xaxis.set_major_locator(mdates.WeekdayLocator(interval=2))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        plt.tight_layout()
        
        # 儲存到BytesIO
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        return img_buffer
        
    def export_to_excel(self, stock_code: str, tables: Dict[str, pd.DataFrame],
                       kline_data: pd.DataFrame, output_file: str):
        """
        輸出到Excel檔案
        
        Args:
            stock_code: 股票代碼
            tables: 三個表格的字典
            kline_data: K線數據
            output_file: 輸出檔案名稱
        """
        wb = Workbook()
        
        # 移除預設的工作表
        wb.remove(wb.active)
        
        # 建立三個工作表
        sheet_configs = [
            ('人數', 'holders', 'Holders Distribution'),
            ('股數', 'shares', 'Shares Distribution'),
            ('占比', 'percentage', 'Percentage Distribution')
        ]
        
        for sheet_name, table_key, chart_title in sheet_configs:
            ws = wb.create_sheet(title=sheet_name)
            
            # 寫入表格數據
            table_data = tables[table_key]
            
            # 寫入標題
            ws.cell(row=1, column=1, value='日期')
            for col_idx, col_name in enumerate(table_data.columns, 2):
                ws.cell(row=1, column=col_idx, value=col_name)
                
            # 寫入數據
            for row_idx, (date, row_data) in enumerate(table_data.iterrows(), 2):
                ws.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
                for col_idx, value in enumerate(row_data, 2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # 建立圖表並插入
            img_buffer = self.create_chart_with_kline(
                table_data, kline_data, 
                f"{stock_code} - {chart_title}",
                sheet_name
            )
            
            img = Image(img_buffer)
            img.width = 900
            img.height = 450
            ws.add_image(img, f'A{len(table_data) + 5}')
            
        # 儲存Excel檔案
        wb.save(output_file)
        logger.info(f"已輸出到 {output_file}")
        
    def run(self, stock_code: str, start_date: str, end_date: str, output_file: Optional[str] = None):
        """
        執行查詢與整理
        
        Args:
            stock_code: 股票代碼
            start_date: 起始日期 (YYYY-MM-DD)
            end_date: 結束日期 (YYYY-MM-DD)
            output_file: 輸出檔案名稱
        """
        logger.info(f"開始查詢股票 {stock_code} 從 {start_date} 到 {end_date}")
        
        # 載入股權分佈數據
        distribution_data = self.load_stock_data(stock_code, start_date, end_date)
        if distribution_data.empty:
            logger.error("無法載入股權分佈數據")
            return
            
        # 獲取K線數據
        kline_data = self.fetch_kline_data(stock_code, start_date, end_date)
        
        # 建立表格
        tables = self.create_tables(distribution_data)
        
        # 輸出到Excel
        if not output_file:
            output_file = f"{stock_code}_{start_date}_{end_date}_analysis.xlsx"
            
        self.export_to_excel(stock_code, tables, kline_data, output_file)
        
def main():
    """主程序"""
    parser = argparse.ArgumentParser(description='股權分佈資料查詢與整理')
    parser.add_argument('stock_code', type=str, help='股票代碼')
    parser.add_argument('start_date', type=str, help='起始日期 (YYYY-MM-DD)')
    parser.add_argument('end_date', type=str, help='結束日期 (YYYY-MM-DD)')
    parser.add_argument('--output', type=str, help='輸出檔案名稱', default=None)
    parser.add_argument('--data-dir', type=str, help='資料目錄', default='stock_data')
    
    args = parser.parse_args()
    
    # 驗證日期格式
    try:
        datetime.strptime(args.start_date, '%Y-%m-%d')
        datetime.strptime(args.end_date, '%Y-%m-%d')
    except ValueError:
        logger.error("日期格式錯誤，請使用 YYYY-MM-DD 格式")
        return
        
    # 執行查詢
    query = StockDataQuery(data_dir=args.data_dir)
    query.run(args.stock_code, args.start_date, args.end_date, args.output)
    
if __name__ == "__main__":
    main()