#!/usr/bin/env python3
"""
程式三：數據分析與繪圖
目的：對程式二的輸出進行深度分析，並根據多種自定義標準繪製趨勢圖
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import logging
import argparse
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class StockAnalysisVisualizer:
    """股權分佈數據分析與視覺化系統"""
    
    def __init__(self):
        """初始化分析系統"""
        # TDCC標準的15個持股級距
        self.standard_levels = [
            '1-999',
            '1,000-5,000',
            '5,001-10,000',
            '10,001-15,000',
            '15,001-20,000',
            '20,001-30,000',
            '30,001-40,000',
            '40,001-50,000',
            '50,001-100,000',
            '100,001-200,000',
            '200,001-400,000',
            '400,001-600,000',
            '600,001-800,000',
            '800,001-1,000,000',
            '1,000,001以上'
        ]
        
    def load_excel_data(self, excel_file: str) -> Dict[str, pd.DataFrame]:
        """
        載入程式二輸出的Excel檔案
        
        Args:
            excel_file: Excel檔案路徑
            
        Returns:
            包含三個表格數據的字典
        """
        try:
            # 讀取Excel檔案的所有工作表
            excel_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
            
            result = {}
            for sheet_name, df in excel_data.items():
                # 設定第一欄為索引（日期）
                if not df.empty:
                    df.set_index(df.columns[0], inplace=True)
                    df.index = pd.to_datetime(df.index)
                    result[sheet_name] = df
                    
            logger.info(f"成功載入 {len(result)} 個工作表")
            return result
            
        except Exception as e:
            logger.error(f"載入Excel檔案失敗: {e}")
            return {}
            
    def parse_level_range(self, level: str) -> Tuple[int, int]:
        """
        解析持股級距字串
        
        Args:
            level: 持股級距字串 (例如: "1,000-5,000")
            
        Returns:
            (最小值, 最大值)
        """
        try:
            # 移除逗號和空格
            level = level.replace(',', '').replace(' ', '')
            
            if '以上' in level or '以上' in level:
                # 處理 "1,000,001以上" 這種格式
                min_val = int(level.replace('以上', '').replace('以上', ''))
                return min_val, float('inf')
            elif '-' in level:
                # 處理 "1,000-5,000" 這種格式
                parts = level.split('-')
                return int(parts[0]), int(parts[1])
            else:
                # 單一數值
                val = int(level)
                return val, val
        except:
            logger.warning(f"無法解析級距: {level}")
            return 0, 0
            
    def categorize_by_shares(self, levels: List[str]) -> Dict[str, List[str]]:
        """
        類別一：一般定義（股數）
        散戶: 1-400,000股
        中實戶: 400,001-1,000,000股
        大戶: 1,000,001+股
        
        Args:
            levels: 持股級距列表
            
        Returns:
            分類後的級距字典
        """
        categories = {
            '散戶': [],
            '中實戶': [],
            '大戶': []
        }
        
        for level in levels:
            min_val, max_val = self.parse_level_range(level)
            
            if max_val <= 400000:
                categories['散戶'].append(level)
            elif min_val <= 1000000 and max_val >= 400001:
                categories['中實戶'].append(level)
            elif min_val > 1000000:
                categories['大戶'].append(level)
                
        return categories
        
    def categorize_by_amount(self, levels: List[str], stock_price: float) -> Dict[str, List[str]]:
        """
        類別二：金額定義（金額 = 股數 x 股價）
        散戶: < 500萬
        小中實戶: 500萬-1,000萬
        中實戶: 1,000萬-3,000萬
        大戶: > 3,000萬
        
        Args:
            levels: 持股級距列表
            stock_price: 股價
            
        Returns:
            分類後的級距字典
        """
        categories = {
            '散戶': [],
            '小中實戶': [],
            '中實戶': [],
            '大戶': []
        }
        
        for level in levels:
            min_val, max_val = self.parse_level_range(level)
            min_amount = min_val * stock_price
            max_amount = max_val * stock_price if max_val != float('inf') else float('inf')
            
            if max_amount <= 5000000:
                categories['散戶'].append(level)
            elif min_amount <= 10000000 and max_amount >= 5000001:
                categories['小中實戶'].append(level)
            elif min_amount <= 30000000 and max_amount >= 10000001:
                categories['中實戶'].append(level)
            elif min_amount > 30000000:
                categories['大戶'].append(level)
                
        return categories
        
    def categorize_custom(self, levels: List[str], custom_ranges: List[Tuple[int, int]]) -> Dict[str, List[str]]:
        """
        類別三：自由手動輸入
        
        Args:
            levels: 持股級距列表
            custom_ranges: 自定義範圍列表 [(min1, max1), (min2, max2), ...]
            
        Returns:
            分類後的級距字典
        """
        categories = {}
        for i, (range_min, range_max) in enumerate(custom_ranges):
            category_name = f"{range_min:,}-{range_max:,}" if range_max != float('inf') else f"{range_min:,}以上"
            categories[category_name] = []
            
            for level in levels:
                min_val, max_val = self.parse_level_range(level)
                
                # 檢查級距是否在範圍內
                if (min_val >= range_min and min_val <= range_max) or \
                   (max_val >= range_min and max_val <= range_max) or \
                   (min_val <= range_min and max_val >= range_max):
                    categories[category_name].append(level)
                    
        return categories
        
    def aggregate_by_category(self, df: pd.DataFrame, categories: Dict[str, List[str]]) -> pd.DataFrame:
        """
        根據分類聚合數據
        
        Args:
            df: 原始數據DataFrame
            categories: 分類字典
            
        Returns:
            聚合後的DataFrame
        """
        aggregated = pd.DataFrame(index=df.index)
        
        for category_name, levels in categories.items():
            # 找出屬於這個類別的所有欄位
            matching_cols = [col for col in df.columns if col in levels]
            if matching_cols:
                # 加總這些欄位
                aggregated[category_name] = df[matching_cols].sum(axis=1)
                
        return aggregated
        
    def create_trend_chart(self, df: pd.DataFrame, title: str, ylabel: str,
                          use_dynamic_scale: bool = True) -> BytesIO:
        """
        建立趨勢圖
        
        Args:
            df: 數據DataFrame
            title: 圖表標題
            ylabel: Y軸標籤
            use_dynamic_scale: 是否使用動態刻度
            
        Returns:
            圖表的BytesIO對象
        """
        # 建立圖表
        fig = Figure(figsize=(16, 10))
        canvas = FigureCanvasAgg(fig)
        ax = fig.add_subplot(111)
        
        # 設定顏色
        colors = plt.cm.tab20(np.linspace(0, 1, len(df.columns)))
        
        # 繪製每條線
        for idx, col in enumerate(df.columns):
            ax.plot(df.index, df[col], label=col, color=colors[idx], linewidth=2, marker='o', markersize=4)
            
        # 設定標籤和標題
        ax.set_xlabel('日期', fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=14, fontweight='bold')
        
        # 設定圖例
        ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=10)
        
        # 格式化X軸日期
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate(rotation=45)
        
        # 動態刻度處理
        if use_dynamic_scale and len(df) > 3:
            for col in df.columns:
                col_data = df[col].dropna()
                if len(col_data) >= 3:
                    # 檢查連續3個數據點的變化範圍
                    for i in range(len(col_data) - 2):
                        window = col_data.iloc[i:i+3]
                        y_range = window.max() - window.min()
                        total_range = col_data.max() - col_data.min()
                        
                        # 如果變化範圍小於總範圍的1%，調整Y軸刻度
                        if total_range > 0 and y_range / total_range < 0.01:
                            # 設定更細緻的Y軸範圍
                            y_min = col_data.min() * 0.95
                            y_max = col_data.max() * 1.05
                            ax.set_ylim(y_min, y_max)
                            logger.info(f"為 {col} 調整Y軸刻度以顯示細微變化")
                            break
                            
        # 添加網格
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # 雙Y軸處理（如果需要）
        if '占比' in ylabel or 'percentage' in ylabel.lower():
            ax2 = ax.twinx()
            ax2.set_ylabel('累積占比 (%)', fontsize=12)
            
            # 計算累積占比
            cumsum_df = df.cumsum(axis=1)
            last_col = cumsum_df.columns[-1]
            ax2.plot(cumsum_df.index, cumsum_df[last_col], 
                    color='red', linewidth=2, linestyle='--', 
                    alpha=0.5, label='累積占比')
            ax2.legend(loc='upper right')
            
        # 調整佈局
        fig.tight_layout()
        
        # 儲存到BytesIO
        img_buffer = BytesIO()
        canvas.print_png(img_buffer)
        img_buffer.seek(0)
        
        return img_buffer
        
    def export_analysis(self, data: Dict[str, pd.DataFrame], categories: Dict[str, List[str]],
                       category_name: str, output_prefix: str):
        """
        輸出分析結果到Excel檔案
        
        Args:
            data: 原始數據字典
            categories: 分類字典
            category_name: 分類名稱
            output_prefix: 輸出檔案前綴
        """
        # 為每個指標建立獨立的Excel檔案
        metrics = [
            ('人數', 'holders', '持股人數'),
            ('股數', 'shares', '持股股數'),
            ('占比', 'percentage', '占集保庫存比例 (%)')
        ]
        
        for sheet_name, metric_key, metric_label in metrics:
            if sheet_name not in data:
                logger.warning(f"找不到 {sheet_name} 工作表，跳過")
                continue
                
            # 建立新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f"{category_name}_{metric_label}"
            
            # 聚合數據
            df = data[sheet_name]
            aggregated_df = self.aggregate_by_category(df, categories)
            
            if aggregated_df.empty:
                logger.warning(f"{sheet_name} 無數據可分析")
                continue
                
            # 寫入數據到工作表
            # 寫入標題
            ws.cell(row=1, column=1, value='日期')
            for col_idx, col_name in enumerate(aggregated_df.columns, 2):
                ws.cell(row=1, column=col_idx, value=col_name)
                
            # 寫入數據
            for row_idx, (date, row_data) in enumerate(aggregated_df.iterrows(), 2):
                ws.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
                for col_idx, value in enumerate(row_data, 2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # 建立趨勢圖
            chart_title = f"{category_name} - {metric_label}趨勢圖"
            img_buffer = self.create_trend_chart(
                aggregated_df, 
                chart_title,
                metric_label
            )
            
            # 插入圖表到工作表
            img = Image(img_buffer)
            img.width = 1200
            img.height = 600
            ws.add_image(img, f'A{len(aggregated_df) + 5}')
            
            # 儲存檔案
            output_file = f"{output_prefix}_{category_name}_{metric_key}.xlsx"
            wb.save(output_file)
            logger.info(f"已輸出分析結果到 {output_file}")
            
    def run(self, input_file: str, stock_price: Optional[float] = None,
            custom_ranges: Optional[List[Tuple[int, int]]] = None):
        """
        執行分析主程序
        
        Args:
            input_file: 程式二輸出的Excel檔案
            stock_price: 股價（用於金額分類）
            custom_ranges: 自定義範圍
        """
        # 載入數據
        data = self.load_excel_data(input_file)
        if not data:
            logger.error("無法載入數據")
            return
            
        # 獲取所有級距
        first_sheet = list(data.values())[0]
        levels = list(first_sheet.columns)
        
        # 從檔名提取股票代碼
        file_stem = Path(input_file).stem
        parts = file_stem.split('_')
        stock_code = parts[0] if parts else 'unknown'
        
        # 執行三種分類分析
        logger.info("執行類別一：一般定義（股數）分析...")
        categories_shares = self.categorize_by_shares(levels)
        self.export_analysis(data, categories_shares, '股數分類', f"{stock_code}_analysis_shares")
        
        if stock_price:
            logger.info(f"執行類別二：金額定義分析（股價={stock_price}）...")
            categories_amount = self.categorize_by_amount(levels, stock_price)
            self.export_analysis(data, categories_amount, '金額分類', f"{stock_code}_analysis_amount")
        else:
            logger.info("未提供股價，跳過金額定義分析")
            
        if custom_ranges:
            logger.info("執行類別三：自定義範圍分析...")
            categories_custom = self.categorize_custom(levels, custom_ranges)
            self.export_analysis(data, categories_custom, '自定義分類', f"{stock_code}_analysis_custom")
        else:
            logger.info("未提供自定義範圍，跳過自定義分析")
            
def parse_custom_ranges(ranges_str: str) -> List[Tuple[int, int]]:
    """
    解析自定義範圍字串
    
    Args:
        ranges_str: 範圍字串，例如 "0-30,30-100,100-500,500+"
        
    Returns:
        範圍列表
    """
    ranges = []
    parts = ranges_str.split(',')
    
    for part in parts:
        part = part.strip().replace(' ', '')
        if '+' in part:
            # 處理 "500+" 這種格式
            min_val = int(part.replace('+', ''))
            ranges.append((min_val, float('inf')))
        elif '-' in part:
            # 處理 "0-30" 這種格式
            min_max = part.split('-')
            ranges.append((int(min_max[0]), int(min_max[1])))
            
    return ranges
    
def main():
    """主程序"""
    parser = argparse.ArgumentParser(description='股權分佈數據分析與視覺化')
    parser.add_argument('input_file', type=str, help='程式二輸出的Excel檔案路徑')
    parser.add_argument('--price', type=float, help='股價（用於金額分類）', default=None)
    parser.add_argument('--custom-ranges', type=str, 
                       help='自定義範圍，例如: "0-30,30-100,100-500,500+"', 
                       default=None)
    
    args = parser.parse_args()
    
    # 解析自定義範圍
    custom_ranges = None
    if args.custom_ranges:
        try:
            custom_ranges = parse_custom_ranges(args.custom_ranges)
            logger.info(f"解析自定義範圍: {custom_ranges}")
        except Exception as e:
            logger.error(f"解析自定義範圍失敗: {e}")
            
    # 執行分析
    analyzer = StockAnalysisVisualizer()
    analyzer.run(args.input_file, args.price, custom_ranges)
    
if __name__ == "__main__":
    main()